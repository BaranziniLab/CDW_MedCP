"""Parse deid_uf_data_dictionary.xlsx into schema_reference.json"""

import json
import sys
from pathlib import Path

import openpyxl

def parse_data_dictionary(xlsx_path: str, output_path: str):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    # Parse Tables sheet
    tables_sheet = wb["Tables"]
    tables_rows = list(tables_sheet.iter_rows(values_only=True))
    tables_header = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(tables_rows[0])]
    tables = {}
    for row in tables_rows[1:]:
        record = dict(zip(tables_header, row))
        name = record.get("table_name")
        if name:
            tables[name] = {
                "description": record.get("table_description", ""),
                "has_patient_data": record.get("has_pat_specific_data") == "Y",
                "has_phi": record.get("has_PHI") == "Y",
                "has_encounter_data": record.get("has_enc_specific_data") == "Y",
                "patient_key_column": record.get("PatientKey_Col"),
                "encounter_key_column": record.get("EncounterKey_Col"),
            }

    # Parse Columns sheet
    cols_sheet = wb["Columns"]
    cols_rows = list(cols_sheet.iter_rows(values_only=True))
    cols_header = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(cols_rows[0])]

    columns_by_table = {}
    for row in cols_rows[1:]:
        record = dict(zip(cols_header, row))
        table_name = record.get("table_name")
        if not table_name:
            continue
        if table_name not in columns_by_table:
            columns_by_table[table_name] = []
        col_info = {
            "name": record.get("column_name"),
            "description": record.get("column_description", ""),
            "data_type": record.get("data_type"),
            "ordinal_position": record.get("ordinal_position"),
        }
        lookup = record.get("lookupTableName")
        if lookup:
            col_info["lookup_table"] = lookup
            col_info["lookup_type"] = record.get("lookupType")
        columns_by_table[table_name].append(col_info)

    # Merge
    schema = {}
    all_table_names = set(tables.keys()) | set(columns_by_table.keys())
    for name in sorted(all_table_names):
        entry = tables.get(name, {"description": "", "has_patient_data": False, "has_phi": False})
        entry["columns"] = columns_by_table.get(name, [])
        schema[name] = entry

    wb.close()

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(schema, f, indent=2, default=str)

    print(f"Wrote {len(schema)} tables to {output_path}")


if __name__ == "__main__":
    project_root = Path(__file__).parent.parent
    xlsx = project_root / "deid_uf_data_dictionary.xlsx"
    output = project_root / "data" / "schema_reference.json"
    parse_data_dictionary(str(xlsx), str(output))
